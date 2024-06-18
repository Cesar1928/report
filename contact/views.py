from django.shortcuts import render, redirect
#from forms_django.wsgi import *
from django.urls import reverse
from .models import Contact1
from django.views.generic import TemplateView, ListView, DetailView, UpdateView, CreateView, DeleteView
from django.http import HttpResponseRedirect
from datetime import *
from xhtml2pdf import pisa
from django import http
from django.template.loader import get_template
from io import BytesIO
import cgi
from django.shortcuts import render
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from PIL import Image as imag
import datetime
from os import remove
import psycopg2 as pg
import pandas.io.sql as psql
from dotenv import load_dotenv
import shutil
import sqlite3
import pandas as pd
import xlsxwriter
import openpyxl 
from openpyxl.drawing.image import Image
from PIL import Image as imag
import datetime
from os import remove
import psycopg2 as pg
import pandas.io.sql as psql
from dotenv import load_dotenv
import shutil
from django.conf import settings
from django.core.mail import  EmailMultiAlternatives, get_connection, send_mail
from django.core.mail.message import EmailMessage
import requests
from .forms import ContactForm
from os import path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import os
from weasyprint import HTML
from datetime import datetime

def Contact(request):
    contact_form = ContactForm()

    if request.method == 'POST':
        contact_form = ContactForm(request.POST, request.FILES)

        if contact_form.is_valid():
            contact_form.save()
            # Tengo que avisar que todo fue bien
            return redirect(reverse('contact')+'?ok')
        
        else:
            #Tengo que generar un error
            return redirect(reverse('contact')+'?error')   

    return render(request, 'contact/contact.html', {'form':contact_form})

# generar el scrip para función (crearpdf)
def crearpdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    HTML(string=html).write_pdf(target="reporte.pdf")
    #f = open(os.path.join(dirname, 'repoe.pdf'), 'wb')
    ruta= path.abspath("reporte.pdf")  
    content = open(ruta, 'rb').read()
    
    subject = 'RAC'
    body = "Líder se envía reporte para atención y tratamiento, saludos "
    attachments = [] 
    attachment = ("reporte.pdf", content, 'application/pdf')
    attachments.append(attachment)
    connection = get_connection(host='smtp.gmail.com', port=587, username="cesarbrianlozada@gmail.com", password=, use_tls=True)
    EmailMessage(subject, body, "cesarbrianlozada@gmail.com", ["clozada@famesa.com.pe"], connection=connection, attachments=attachments).send()       
    connection.close()

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = http.HttpResponse(result.getvalue(), content_type= "application/pdf")
        conte = "attachment; filename = {0}".format("reporte.pdf")
        response["Content-Disposition"] = conte
              
        #***********************otro ejemplo
        #pdf = HTML(string=html).write_pdf(response)
        #subject = 'The Subject of the mail'
        #body = "jaa"
        
        #attachments = []  # start with an empty list
        #ruta= path.abspath("a.pdf")  
        #content = open("C:/Users/WIN-10/Desktop/PYTHON-DJANGO/forms_django/a.pdf", 'rb').read()
        #attachment = ("reporte.pdf", content, 'application/pdf')
        #attachments.append(attachment)
        #connection = get_connection(host='smtp.gmail.com', port=587, username="cesarbrianlozada@gmail.com", password=, use_tls=True)
        #EmailMessage(subject, body, "cesarbrianlozada@gmail.com", ["clozada@famesa.com.pe"], connection=connection, attachments=attachments).send()       
        #connection.close()

        return response
    return http.HttpResponse('Ocurrio un error al genera el reporte %s' % cgi.escape(html))

"""
def enviarcorreo (request):
    contact_form = ContactForm(request.POST, request.FILES)
    template = get_template("contact/generacion_reporte.html")
    html = template.render({'form':contact_form})
    pdf = HTML(string=html).write_pdf(target="re19.pdf")
    #f = open(os.path.join(dirname, 'repoe.pdf'), 'wb')
    ruta= path.abspath("re1.pdf")  
    content = open(ruta, 'rb').read()

    subject = 'The Subject of the mail'
    body = "jaa"
    attachments = [] 
    attachment = ("re19.pdf", content, 'application/pdf')
    attachments.append(attachment)


"""
"""
def enviarcorreo(request):
    contact_form = ContactForm()
    subject = 'The Subject of the mail'
    body = "jaa"

    
    return render(request, 'contact/contact.html', {'form':contact_form})
"""

#aplicación de la función anterior 
def generar_reporte_productos(request):
    contact = Contact1.objects.all()
    for expe in contact:
        nombre = expe.name
        lugar = expe.lugar
        subestandar = expe.subestandar
        empresa = expe.empresa
        hora = expe.hora
        fecha = expe.fecha
        area = expe.area
        descripcion = expe.descripcion
        riesgo = expe.riesgo
        imagen = expe.imagen
        accion = expe.accion
        recomendacion = expe.recomendacion
        codigo = expe.codigo
        areareportante = expe.areareportante

    dominio = "http://localhost:8000/media/"
    dominio1 = dominio + str(imagen)
    dominio1 = str(dominio1)

    riesgo1=[]
    riesgo2=[]
    riesgo3=[]
    if riesgo == "Bajo":
        riesgo1 = "( )"
        riesgo2 = "( )"
        riesgo3 = "(X)"
    elif riesgo=="Medio":
        riesgo1 = "( )"
        riesgo2 = "(X)"
        riesgo3 = "( )"
    else:
        riesgo1 = "(X)"
        riesgo2 = "( )"
        riesgo3 = "( )"


    subestandar2=[]
    subestandar1=[]
    if subestandar == "Condición Sub-estándar":
        subestandar2 = "( )"
        subestandar1 = "(X)"
    else:
        subestandar1 = "( )"
        subestandar2 ="(X)"

    empresa2=[]
    empresa1=[]
    if empresa == "Famesa Explosivos":
        empresa1 = "( )"
        empresa2 = "(X)"
    else:
        empresa2 = "( )"
        empresa1 ="(X)"

    crearp = crearpdf('contact/generacion_reporte.html',{'nombre': nombre,
                                                        "subestandar1":subestandar1,
                                                        "subestandar2":subestandar2,
                                                        "empresa1":empresa1,
                                                        "empresa2":empresa2,
                                                        "riesgo1":riesgo1,
                                                        "riesgo2":riesgo2,
                                                        "riesgo3":riesgo3,
                                                       "lugar":lugar,  
                                                       "hora":hora,
                                                       "fecha":fecha,
                                                       "area":area,
                                                       "descripcion":descripcion, 
                                                       "imagen":imagen,
                                                        "accion" : accion,
                                                        "recomendacion" : recomendacion,
                                                        "codigo" : codigo,
                                                        "areareportante" : areareportante,    
                                                        "dominio1":dominio1                                            
                                                       
                                                       }) 
    return crearp

"""
def enviarcorreo(request):
    subject = 'The Subject of the mail'
    body = "holaaaaaaaa"
    connection = get_connection(host='smtp.gmail.com', port=587, username="cesarbrianlozada@gmail.com", password="xdmq dsql gdoe hqtj", use_tls=True)
    EmailMessage(subject, body, "cesarbrianlozada@gmail.com", ["clozada@famesa.com.pe"], connection=connection).send()
    connection.close()
    with open("reporte.pdf", "rb") as f:
    email.add_attachment(
        f.read(),
        filename="reporte.pdf",
        maintype="application",
        subtype="pdf"
    )
    return redirect ("contact/contact.html)
"""
"""
def emailMinutesMeeting(to_email, filename, pdf_location):
    from_email = settings.EMAIL_HOST_USER

    subject = 'Subject of the email'
    text_content = "hola"

    msg = EmailMultiAlternatives(subject, text_content, from_email, [to_email])
    msg.attach_alternative(html_content, 'text/html')

    file_data = open(pdf_location, 'rb')
    msg.attach(filename, file_data.read(), "application/pdf")

    file_data.close()
    msg.send()
    return redirect ("contact/contact.html)
    return render(request, 'contact/contact.html')

""" 

class ReportePersonalizadoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        remove("rac1.xlsx") #el rac final de reporte final
        shutil.copy("cop.xlsx", "rac1.xlsx")
        conn = pg.connect("dbname=rac1 user=postgres password=cesar")
        cursor = conn.cursor()
        query = "SELECT * FROM contact_contact1"
        cursor.execute(query)
        # The execute returns a list of tuples:
        tuples_list = cursor.fetchall()
        column_names = ["id",'name',"lugar","fecha","area","hora","subestandar","empresa","riesgo","descripcion", "accion","recomendacion","codigo","areareportante","imagen"]
            # Now we need to transform the list into a pandas DataFrame:
        df = pd.DataFrame(tuples_list, columns=column_names)
        df["imagen"] = "media/" + df["imagen"].astype(str) 
        b = 14
        for i in range(len(df)):  
            b = b + 1
            imagen = imag.open(df.loc[i, "imagen"])
            imagen = imagen.convert('RGBA')
            ancho, alto = imagen.size
            imagen = imagen.resize((170, 70))
            imagen.save("tunel77.png")
            imagen1 = Image("tunel77.png")
            a = "N"+str(b)
            #e = "rac"+str(d)+".xlsx"  # el rac de base 
            wb = openpyxl.load_workbook("rac1.xlsx") 
            ws = wb.active
            ws.column_dimensions["O"].width = 30
            ws.add_image(imagen1, a) 
            ws.cell(row=b, column=5, value = df.loc[i,"name"]) #para referiar contenido texto o númericos
            ws.cell(row=b, column=2, value = df.loc[i,"id"])
            ws.cell(row=b, column=4, value = df.loc[i,"descripcion"])
            ws.cell(row=b, column=8, value = df.loc[i,"riesgo"])
            ws.cell(row=b, column=9, value = df.loc[i,"accion"])
            ws.cell(row=10, column=4, value = "Asistencia Técnica")
            ws.cell(row=10, column=8, value = "Ing. Fredy Herrera")
            ws.cell(row=b, column=7, value = "Correctiva")
            ws.cell(row=b, column=3, value = "Inspección Inopinada")
            ws.cell(row=b, column=6, value = "Asistencia Técnica")
            ws.cell(row=b, column=10, value = "Alvino Fabian")
            fechahoy = datetime.date.today()
            ws.cell(row=10, column=12, value = fechahoy)
            #wb.save('rac1.xlsx')
            #wb.close()
            #Establecer el nombre de mi archivo
            #wb.save("rac1.xlsx")
        nombre_archivo = "ra.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
"""
class Send(View):
    def get(self, request):
        return render(request, 'mail/send.html')
    
    def post(self, request):
        email = request.POST.get('email')
        print(email)

        template = get_template('mail/email-order-success.html')

        # Se renderiza el template y se envias parametros
        content = template.render({'email': email})

        # Se crea el correo (titulo, mensaje, emisor, destinatario)
        msg = EmailMultiAlternatives(
            'Gracias por tu compra',
            'Hola, te enviamos un correo con tu factura',
            settings.EMAIL_HOST_USER,
            [email]
        )

        msg.attach_alternative(content, 'text/html')
        msg.send()

        return render(request, 'mail/send.html')

"""


"""

#Nuestra clase hereda de la vista genérica TemplateView
class ReportePersonasExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        
        conn = pg.connect("dbname=rac1 user=postgres password=cesar")
        cursor = conn.cursor()
        query = "SELECT * FROM contact_contact1"
        cursor.execute(query)
        # The execute returns a list of tuples:
        tuples_list = cursor.fetchall()
        column_names = ["id",'name',"lugar","fecha","area","hora","subestandar","empresa","riesgo","descripcion", "accion","recomendacion","codigo","areareportante","imagen"]
            # Now we need to transform the list into a pandas DataFrame:
        df = pd.DataFrame(tuples_list, columns=column_names)
        df["imagen"] = "media/" + df["imagen"].astype(str) 

        b = 14
        for i in range(len(df)):  
            b = b + 1
            imagen = imag.open(df.loc[i, "imagen"])
            rgb_im = imagen.convert('RGBA')
            ancho, alto = imagen.size
            imagen = imagen.resize((170, 80))
            imagen.save("tunel777.jpg")
            imagen1 = Image("tunel777.jpg")
            a = "N"+str(b)
            #e = "rac"+str(d)+".xlsx"  # el rac de base 
            wb = openpyxl.load_workbook("rac1.xlsx") 
            ws = wb.active
            ws.column_dimensions["O"].width = 30
            ws.add_image(imagen1, a) 
            ws.cell(row=b, column=5, value = df.loc[i,"name"]) #para referiar contenido texto o númericos
            ws.cell(row=b, column=2, value = df.loc[i,"id"])
            ws.cell(row=b, column=4, value = df.loc[i,"descripcion"])
            ws.cell(row=b, column=8, value = df.loc[i,"riesgo"])
            ws.cell(row=b, column=9, value = df.loc[i,"accion"])
            ws.cell(row=10, column=4, value = "Asistencia Técnica")
            ws.cell(row=10, column=8, value = "Ing. Fredy Herrera")
            ws.cell(row=b, column=7, value = "Correctiva")
            ws.cell(row=b, column=3, value = "Inspección Inopinada")
            ws.cell(row=b, column=6, value = "Asistencia Técnica")
            ws.cell(row=b, column=10, value = "Alvino Fabian")
            fechahoy = datetime.date.today()
            ws.cell(row=10, column=12, value = fechahoy)
            nombre_archivo ="ReportePersonasExcel.xlsx"
            #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
            response = HttpResponse(content_type="application/ms-excel") 
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

"""