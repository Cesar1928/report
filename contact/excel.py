import sqlite3
import pandas as pd
import xlsxwriter
import openpyxl 
from openpyxl.drawing.image import Image
from PIL import Image as imag
import datetime
from os import remove
import psycopg2 as pg
import pandas.io.sql as psql
from dotenv import load_dotenv
import shutil
 
#con = sqlite3.connect ("db.sqlite3")
#cursor = con.cursor()
#sql_query = pd.read_sql("SELECT * FROM contact_contact1", con)
#df = pd.DataFrame(sql_query, columns = ["id",'name',"codigo","areareportante","lugar","fecha","area","hora","subestandar","empresa","riesgo","descripcion", "accion","recomendacion","imagen"])

conn = pg.connect("dbname=rac1 user=postgres password=cesar")
cursor = conn.cursor()
query = "SELECT * FROM contact_contact1"
cursor.execute(query)
# The execute returns a list of tuples:
tuples_list = cursor.fetchall()
column_names = ["id",'name',"lugar","fecha","area","hora","subestandar","empresa","riesgo","descripcion", "accion","recomendacion","codigo","areareportante","imagen"]
    # Now we need to transform the list into a pandas DataFrame:
df = pd.DataFrame(tuples_list, columns=column_names)
df["imagen"] = "media/" + df["imagen"].astype(str) 
print(df)

#verificar siempre el rac1 en la carpeta
remove("rac1.xlsx") #el rac final de reporte final
shutil.copy("cop.xlsx", "rac1.xlsx")

b = 14
for i in range(len(df)):  
    b = b + 1
    imagen = imag.open(df.loc[i, "imagen"])
    rgb_im = imagen.convert('RGBA')
    ancho, alto = imagen.size
    imagen = imagen.resize((170, 80))
    imagen.save("tunel77.png")
    imagen1 = Image("tunel77.png")
    a = "N"+str(b)
    #e = "rac"+str(d)+".xlsx"  # el rac de base 
    wb = openpyxl.load_workbook("rac1.xlsx") 
    ws = wb.active
    ws.column_dimensions["O"].width = 30
    ws.add_image(imagen1, a) 
    ws.cell(row=b, column=5, value = df.loc[i,"name"]) #para referiar contenido texto o númericos
    ws.cell(row=b, column=2, value = df.loc[i,"id"])
    ws.cell(row=b, column=4, value = df.loc[i,"descripcion"])
    ws.cell(row=b, column=8, value = df.loc[i,"riesgo"])
    ws.cell(row=b, column=9, value = df.loc[i,"accion"])
    ws.cell(row=10, column=4, value = "Asistencia Técnica")
    ws.cell(row=10, column=8, value = "Ing. Fredy Herrera")
    ws.cell(row=b, column=7, value = "Correctiva")
    ws.cell(row=b, column=3, value = "Inspección Inopinada")
    ws.cell(row=b, column=6, value = "Asistencia Técnica")
    ws.cell(row=b, column=10, value = "Alvino Fabian")
    fechahoy = datetime.date.today()
    ws.cell(row=10, column=12, value = fechahoy)
    wb.save('rac1.xlsx')
print("SE EXPORTÓ EL EXCEL - CONSOLIDADO DE REPORTES OCURRENCIAS")

#a = 14
#for i in df["name"]:
#    a = a+1
#    ws.cell(row=a, column=7, value = i)
#wb.save('rac1.xlsx')




# leemos el fichero
#libro = load_workbook('fichero.xlsx')
# primera pestaña o pestaña activa
#hoja = libro.active
# mostramos la celda B1 = 1:2 = Fila 1 - Columna B (2)
#print(hoja.cell(row=1, column=2).value)
# mostramos la celda B1 = 1:2 = Fila 1 - Columna B (2)
#print(hoja['B1'].value)
#ws.cell(row=15, column=7, value = df["name"].iloc[-1])
#valordecelda = ws['A1'].value
#https://www.datacamp.com/es/tutorial/python-excel-tutorial
